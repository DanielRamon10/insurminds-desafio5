"""Monta a planilha de trabalho do especialista em seguros (frente B).

O objetivo é que a contribuição de domínio — quais números caracterizam cada
evento e quando ele vira alerta — chegue sem exigir Git, terminal ou Python.
A planilha sai pronta para preencher no Excel e voltar por e-mail.

Uso:
    python scripts/gerar_planilha_ph.py
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parent.parent
DESTINO = RAIZ / "docs" / "FRENTE_B_limiares_e_segurados.xlsx"

TINTA = "0F1F2B"
ACENTO = "0A6B8A"
CLARO = "E4F0F5"
AMBAR = "FDF3E3"
BORDA = Side(style="thin", color="C9D6DF")


def _titulo(ws, linha: int, texto: str, largura: int) -> None:
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura)
    c = ws.cell(row=linha, column=1, value=texto)
    c.font = Font(bold=True, size=13, color=TINTA)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[linha].height = 22


def _nota(ws, linha: int, texto: str, largura: int = 7) -> None:
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura)
    c = ws.cell(row=linha, column=1, value=texto)
    c.font = Font(size=10, color="3D5261")
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[linha].height = 46


def _cabecalho(ws, linha: int, colunas: list[str]) -> None:
    for i, nome in enumerate(colunas, start=1):
        c = ws.cell(row=linha, column=i, value=nome)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=ACENTO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=BORDA)
    ws.row_dimensions[linha].height = 30


def aba_limiares(wb: Workbook) -> None:
    """A entrega principal: os números que só o especialista sabe definir."""
    ws = wb.create_sheet("1. Limiares (PREENCHER)")
    ws.sheet_properties.tabColor = ACENTO

    _titulo(ws, 1, "Limiares dos eventos climáticos", 7)
    _nota(ws, 2,
          "Preencha as colunas amarelas. A pergunta é: a partir de que número o evento "
          "passa a merecer um aviso ao segurado, e a partir de que número ele vira alerta? "
          "Se discordar da coluna de referência, sobrescreva — ela é só ponto de partida. "
          "A coluna 'Por quê' é a mais importante: ela vai direto para o relatório.")

    colunas = ["Evento", "O que medimos", "Unidade",
               "Referência inicial", "ATENÇÃO a partir de", "ALERTA a partir de",
               "Por quê / fonte"]
    _cabecalho(ws, 4, colunas)

    linhas = [
        ("Chuva intensa", "Chuva acumulada em 24h", "mm", "40 mm", "", "", ""),
        ("Chuva intensa", "Chuva na hora mais forte", "mm/h", "10 mm/h", "", "", ""),
        ("Vento forte", "Rajada máxima", "km/h", "60 km/h", "", "", ""),
        ("Raio", "Energia da tempestade (CAPE)", "J/kg", "800 J/kg", "", "", ""),
        ("Granizo", "Energia da tempestade (CAPE)", "J/kg", "1.500 J/kg", "", "", ""),
        ("Granizo", "Altura onde congela", "metros", "abaixo de 3.800 m", "", "", ""),
    ]
    for i, dados in enumerate(linhas, start=5):
        for j, valor in enumerate(dados, start=1):
            c = ws.cell(row=i, column=j, value=valor)
            c.border = Border(bottom=BORDA, left=BORDA, right=BORDA)
            c.alignment = Alignment(vertical="center", wrap_text=(j == 7))
            if j in (5, 6, 7):
                c.fill = PatternFill("solid", fgColor=AMBAR)
        ws.row_dimensions[i].height = 30

    for col, larg in zip("ABCDEFG", [18, 26, 10, 20, 20, 20, 46]):
        ws.column_dimensions[col].width = larg

    _titulo(ws, 13, "Recomendação preventiva por cenário", 4)
    _nota(ws, 14,
          "O que o segurado deve fazer ao receber o aviso. Uma frase por linha, "
          "em linguagem de cliente. É isto que a inteligência artificial vai usar "
          "para escrever a mensagem — quanto mais concreta a orientação, melhor a mensagem.")

    _cabecalho(ws, 16, ["Evento", "Tipo de apólice", "O que o segurado deve fazer", "Observações"])
    cenarios = [
        ("Chuva intensa", "Residencial", "", ""),
        ("Chuva intensa", "Automotiva", "", ""),
        ("Raio", "Residencial", "", ""),
        ("Vento forte", "Residencial", "", ""),
        ("Vento forte", "Automotiva", "", ""),
        ("Granizo", "Residencial", "", ""),
        ("Granizo", "Automotiva", "", ""),
    ]
    for i, dados in enumerate(cenarios, start=17):
        for j, valor in enumerate(dados, start=1):
            c = ws.cell(row=i, column=j, value=valor)
            c.border = Border(bottom=BORDA, left=BORDA, right=BORDA)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if j in (3, 4):
                c.fill = PatternFill("solid", fgColor=AMBAR)
        ws.row_dimensions[i].height = 32

    ws.cell(row=25, column=1,
            value="Observação: raio + apólice automotiva não entra na lista de propósito — "
                  "um carro é uma gaiola de Faraday e protege quem está dentro. "
                  "Se discordar, escreva aqui.").font = Font(size=9, italic=True, color="8D5006")


def aba_segurados(wb: Workbook) -> None:
    """Base já gerada: o especialista revisa em vez de digitar 45 linhas."""
    ws = wb.create_sheet("2. Segurados (revisar)")

    _titulo(ws, 1, "Base de segurados fictícios — já gerada", 8)
    _nota(ws, 2,
          "Estes 45 segurados já estão no sistema. Você não precisa digitar nada aqui: "
          "só confira se a distribuição faz sentido para um cenário de seguradora. "
          "Se algo estiver estranho — proporção de apólices, canais, concentração em "
          "alguma região — anote na última coluna que a gente ajusta.")

    origem = RAIZ / "data" / "segurados.csv"
    if not origem.is_file():
        raise FileNotFoundError(f"Rode antes: python scripts/gerar_segurados.py ({origem})")

    with origem.open(encoding="utf-8", newline="") as f:
        linhas = list(csv.reader(f))

    _cabecalho(ws, 4, [c.replace("_", " ").title() for c in linhas[0]] + ["Comentário"])
    for i, linha in enumerate(linhas[1:], start=5):
        for j, valor in enumerate(linha, start=1):
            c = ws.cell(row=i, column=j, value=valor)
            c.border = Border(bottom=BORDA, left=BORDA, right=BORDA)
        ws.cell(row=i, column=9).fill = PatternFill("solid", fgColor=AMBAR)
        ws.cell(row=i, column=9).border = Border(bottom=BORDA, left=BORDA, right=BORDA)

    for col, larg in zip("ABCDEFGHI", [10, 26, 14, 16, 6, 11, 11, 9, 30]):
        ws.column_dimensions[col].width = larg
    ws.freeze_panes = "A5"


def aba_referencia(wb: Workbook) -> None:
    ws = wb.create_sheet("3. Referência")

    _titulo(ws, 1, "Como o sistema funciona, em quatro passos", 5)
    _nota(ws, 2,
          "Contexto para preencher a aba 1. O sistema consulta a previsão do tempo de "
          "15 cidades, identifica se algum evento vai acontecer, decide quais segurados "
          "avisar e escreve a mensagem. Seus limiares definem o segundo e o terceiro passo.")

    passos = [
        ("1. Consultar", "Busca a previsão das próximas 24h para 15 cidades"),
        ("2. Identificar", "Compara os números com SEUS limiares e nomeia o evento"),
        ("3. Decidir", "Cruza o evento com o tipo de apólice de cada segurado"),
        ("4. Escrever", "A IA redige a mensagem usando SUA recomendação preventiva"),
    ]
    _cabecalho(ws, 4, ["Passo", "O que acontece"])
    for i, (a, b) in enumerate(passos, start=5):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True)
        for j in (1, 2):
            ws.cell(row=i, column=j).border = Border(bottom=BORDA, left=BORDA, right=BORDA)
        ws.row_dimensions[i].height = 24

    _titulo(ws, 11, "As 15 cidades monitoradas", 5)
    _cabecalho(ws, 12, ["Cidade", "UF", "Região climática", "Por que está na lista"])
    porques = {
        "sul": "granizo e frentes frias",
        "litoral": "vento forte",
        "sudeste": "chuva intensa",
        "centro-oeste": "tempestades de verão",
        "norte": "descargas elétricas",
    }
    with (RAIZ / "data" / "cidades.csv").open(encoding="utf-8", newline="") as f:
        for i, linha in enumerate(csv.DictReader(f), start=13):
            reg = linha["regiao_climatica"]
            for j, valor in enumerate(
                [linha["nome"], linha["uf"], reg, porques.get(reg, "")], start=1
            ):
                c = ws.cell(row=i, column=j, value=valor)
                c.border = Border(bottom=BORDA, left=BORDA, right=BORDA)

    for col, larg in zip("ABCD", [20, 8, 18, 30]):
        ws.column_dimensions[col].width = larg


def main() -> int:
    wb = Workbook()
    wb.remove(wb.active)
    aba_limiares(wb)
    aba_segurados(wb)
    aba_referencia(wb)
    DESTINO.parent.mkdir(exist_ok=True)
    wb.save(DESTINO)
    print(f"Planilha gerada: {DESTINO.relative_to(RAIZ)}")
    print(f"Abas: {', '.join(wb.sheetnames)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
